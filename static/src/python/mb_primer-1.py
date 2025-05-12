#! /usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import statistics

wb = load_workbook(filename='mb_primer.xlsx', read_only=True)
ws = wb['Sheet1']

# zahtevana marka betona
MB = ws["I14"].value

# Podatki o preiskušancih
cells = ws["C33":"C52"]
x = []
for c in cells:
	x.append(c[0].value)

# srednja vrednost
mn = sum(x) / len(x) 
print("mn = {:.2f}".format(mn))

# standardni odklon
sn = statistics.stdev(x)
print("sn = {:.4f}".format(sn))

# 1. kriterij
if mn > MB + 1.3 * sn:
	print("mn > MB + 1.3 * sn ... kriterij je izpolnjen")
else:
	print("mn < MB + 1.3 * sn ... kriterij NI izpolnjen")

# 2. kriterij
if min(x) > MB - 4:
	print("xmin > MB - 4 ... kriterij je izpolnjen")
else:
	print("xmin > MB - 4 ... kriterij NI izpolnjen")
